"""
Extrae el detalle por seccion presupuestal del Proyecto de Ley PGN 2027 desde
el Excel oficial y lo guarda en data/raw/pgn2027_raw.json.

Fuente: "Presupuesto General de la Nación..xlsx" (la version de $634,95
billones, articulado radicado agosto 2026 -- ver CLAUDE.md para el porque de
este archivo y no el otro).

Requiere: pip install openpyxl
Uso:      python scripts/1_extract_2027_xlsx.py
"""
import json
import os
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC_XLSX = os.path.join(ROOT, "Presupuesto General de la Nación..xlsx")
DEST = os.path.join(ROOT, "data", "raw", "pgn2027_raw.json")

# Mapa de columnas del Excel (hoja "PGN"), confirmado celda por celda:
#   C = codigo de programa (CTA PROG) O etiqueta "A./B./C." O "TOTAL PRESUPUESTO SECCION"
#   D = codigo de subprograma (SUBC SUBP)
#   E = nombre de programa/subprograma
#   F = "SECCION: NNNN" (fila N) y nombre de la seccion (fila N+1)
#   G = APORTE NACIONAL (valor)
#   H = RECURSOS PROPIOS (valor)
#   I = TOTAL (valor)


def num(v):
    if v is None or v == "" or isinstance(v, str):
        return 0.0
    return float(v)


def main():
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    ws = wb["PGN"]

    sections = []
    cur = None
    r, maxr = 1, ws.max_row

    while r <= maxr:
        c = ws[f"C{r}"].value
        d = ws[f"D{r}"].value
        e = ws[f"E{r}"].value
        f = ws[f"F{r}"].value
        g, h, i = num(ws[f"G{r}"].value), num(ws[f"H{r}"].value), num(ws[f"I{r}"].value)

        if isinstance(f, str) and f.strip().startswith("SECCIÓN:"):
            code = f.split(":")[1].strip()
            name = (ws[f"F{r+1}"].value or "").strip()
            cur = {
                "code": code, "name": name,
                "funcionamiento": 0.0, "deuda": 0.0, "inversion": 0.0,
                "aporte_nacional": 0.0, "recursos_propios": 0.0, "total": 0.0,
                "programas": [],
            }
            sections.append(cur)
            r += 2
            continue

        if cur is None:
            r += 1
            continue

        if isinstance(c, str) and c.strip().startswith("A.") and "FUNCIONAMIENTO" in c.upper():
            cur["funcionamiento"] += i
        elif isinstance(c, str) and c.strip().startswith("B.") and "DEUDA" in c.upper():
            cur["deuda"] += i
        elif isinstance(c, str) and c.strip().startswith("C.") and "INVERSIÓN" in c.upper():
            cur["inversion"] += i
        elif isinstance(c, str) and c.strip() == "TOTAL PRESUPUESTO SECCIÓN":
            cur["aporte_nacional"] += g
            cur["recursos_propios"] += h
            cur["total"] += i
            cur = None
        elif isinstance(c, str) and c.strip() != "" and (d is None or str(d).strip() == "") and isinstance(e, str) and e.strip() != "":
            cur["programas"].append({
                "codigo": c.strip(), "nombre": e.strip(),
                "aporte_nacional": g, "recursos_propios": h, "total": i,
                "subprogramas": [],
            })
        elif (c is None or str(c).strip() == "") and isinstance(d, str) and d.strip() != "" and isinstance(e, str) and e.strip() != "":
            if cur["programas"]:
                cur["programas"][-1]["subprogramas"].append({
                    "codigo": d.strip(), "nombre": e.strip(),
                    "aporte_nacional": g, "recursos_propios": h, "total": i,
                })
        r += 1

    total = sum(s["total"] for s in sections)
    print(f"Secciones parseadas: {len(sections)}")
    print(f"Total (debe ser 634952040871099): {total:.0f}")
    assert abs(total - 634952040871099) < 1000, "El total no coincide con el esperado; revisar el mapeo de columnas."

    os.makedirs(os.path.dirname(DEST), exist_ok=True)
    with open(DEST, "w", encoding="utf-8") as fh:
        json.dump(sections, fh, ensure_ascii=False, indent=1)
    print("Escrito:", DEST)


if __name__ == "__main__":
    main()
